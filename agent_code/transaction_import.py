"""
Parse CSV / Excel uploads into rows for daily_transactions.
Expected columns (headers are matched case-insensitively):
  date / transaction_date, type, category, amount, description (optional)
"""
from __future__ import annotations

import csv
import io
import re
from datetime import date, datetime, timedelta
from decimal import Decimal, InvalidOperation
from typing import Any

logger = None  # set from app if needed


def _norm_header(h: str) -> str:
    return re.sub(r"[^a-z0-9]+", "_", (h or "").strip().lower()).strip("_")


def _find_col(headers: list[str], candidates: tuple[str, ...]) -> int | None:
    norm = [_norm_header(h) for h in headers]
    for cand in candidates:
        c = _norm_header(cand)
        if c in norm:
            return norm.index(c)
    return None


def _parse_date(val: Any):
    if val is None or val == "":
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    s = str(val).strip()
    if not s:
        return None
    if len(s) >= 10 and s[4] == "-" and s[7] == "-":
        try:
            return datetime.strptime(s[:10], "%Y-%m-%d").date()
        except ValueError:
            pass
    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(s[:10], fmt).date()
        except ValueError:
            continue
    # Excel serial (approximate)
    try:
        n = float(s.replace(",", ""))
        if 20000 < n < 60000:
            base = datetime(1899, 12, 30)
            return (base + timedelta(days=int(n))).date()
    except (ValueError, OverflowError):
        pass
    return None


def _parse_amount(val: Any) -> float | None:
    if val is None or val == "":
        return None
    if isinstance(val, (int, float)):
        return float(val)
    s = re.sub(r"[₹$,]", "", str(val).strip())
    if not s:
        return None
    try:
        return float(Decimal(s))
    except (InvalidOperation, ValueError):
        return None


def _parse_type(val: Any) -> str | None:
    if val is None:
        return None
    s = str(val).strip().lower()
    if not s:
        return None
    if s in ("revenue", "income", "sale", "sales", "credit", "cr"):
        return "Revenue"
    if s in ("expense", "cost", "debit", "dr", "payment"):
        return "Expense"
    if "rev" in s or "income" in s or "sale" in s:
        return "Revenue"
    if "exp" in s or "cost" in s:
        return "Expense"
    return None


def _rows_from_dicts(
    headers: list[str], data_rows: list[list[Any]]
) -> list[tuple]:
    """Returns list of (transaction_date, type, category, amount, description)."""
    i_date = _find_col(
        headers,
        ("date", "transaction_date", "txn_date", "day", "posted_date"),
    )
    i_type = _find_col(headers, ("type", "txn_type", "transaction_type", "dr_cr"))
    i_cat = _find_col(headers, ("category", "cat", "account", "head"))
    i_amt = _find_col(headers, ("amount", "value", "total", "amt"))
    i_desc = _find_col(headers, ("description", "desc", "notes", "memo", "particulars"))

    if i_date is None or i_amt is None:
        raise ValueError(
            "Could not find required columns. Include a date column and an amount column "
            "(e.g. date, amount)."
        )

    out: list[tuple] = []
    for row in data_rows:
        if not row or all(
            c is None or str(c).strip() == "" for c in row
        ):
            continue
        def get(i: int | None) -> Any:
            if i is None or i >= len(row):
                return None
            return row[i]

        d = _parse_date(get(i_date))
        amt = _parse_amount(get(i_amt))
        if d is None or amt is None:
            continue

        t = _parse_type(get(i_type)) if i_type is not None else None
        if t is None:
            t = "Revenue" if amt >= 0 else "Expense"
        amt = abs(float(amt))

        cat = str(get(i_cat) or "General").strip() or "General"
        desc = str(get(i_desc) or "").strip() or cat

        out.append((d, t, cat[:100], amt, desc[:500]))

    if not out:
        raise ValueError("No valid data rows found. Check date and amount formats.")
    return out


def parse_csv_bytes(raw: bytes) -> list[tuple]:
    text = raw.decode("utf-8-sig", errors="replace")
    reader = csv.reader(io.StringIO(text))
    rows = list(reader)
    if len(rows) < 2:
        raise ValueError("CSV must have a header row and at least one data row.")
    headers = [str(c or "") for c in rows[0]]
    data_rows = [[c for c in r] for r in rows[1:]]
    return _rows_from_dicts(headers, data_rows)


def parse_xlsx_bytes(raw: bytes) -> list[tuple]:
    from openpyxl import load_workbook

    bio = io.BytesIO(raw)
    wb = load_workbook(bio, read_only=True, data_only=True)
    try:
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        first = next(rows_iter, None)
        if not first:
            raise ValueError("Excel sheet is empty.")
        headers = [str(c) if c is not None else "" for c in first]
        data_rows = []
        for row in rows_iter:
            data_rows.append([c for c in row])
        return _rows_from_dicts(headers, data_rows)
    finally:
        wb.close()